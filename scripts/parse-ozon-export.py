import openpyxl
import re
import json
import os

def parse_size(name):
    # Try to find format like 183х92х82см or 120x60 (with different x letters)
    # Looking for sequences of numbers separated by x or х
    name = name.lower().replace(" ", "")
    # matching 100x50x30 or 100x50
    match = re.search(r'(\d+)[xх](\d+)(?:[xх](\d+))?', name)
    if match:
        dims = [int(x) for x in match.groups() if x is not None]
        # Assuming the largest dimension is length
        length = max(dims)
        size_text = "х".join(str(d) for d in dims)
        return size_text, length
    return None, None

def get_feet_category(length):
    if length is None:
        return "Unknown"
    feet = round(length / 30.48)
    if feet < 3:
        return "-3ft"
    elif feet > 7:
        return "7ft" # Limit to 7ft based on columns, or keep original? Let's cap at 7 or keep going. The columns are 7ft.
    else:
        return f"{feet}ft"

def safe_float(val):
    try:
        if val is None or str(val).strip() == '':
            return 0.0
        return float(str(val).replace(',', '.'))
    except ValueError:
        return 0.0

def safe_int(val):
    try:
        if val is None or str(val).strip() == '':
            return 0
        return int(float(str(val).replace(',', '.')))
    except ValueError:
        return 0

def process_file():
    input_file = r'd:\!PROJECTS\CONCURENTS\MANAGER\ТЗ на обработку выгрузки.xlsx'
    output_file = r'd:\!PROJECTS\CONCURENTS\app\src\data\ozon\products-export.json'
    
    # ensure output dir exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    sheet = wb.active
    
    products = []
    
    # Headers are at row 5
    # Data starts at row 7
    for row in range(7, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        if not name:
            continue
            
        name = str(name)
        # Filter exclusions
        name_lower = name.lower()
        if "футбол" in name_lower or "трансформер" in name_lower or ("хоккей" in name_lower and "аэро" not in name_lower and "настольная игра" not in name_lower):
            if not ("аэрохоккей" in name_lower):
                continue
        
        url = sheet.cell(row=row, column=2).value or ""
        
        # Extract ID from url
        product_id = ""
        if url:
            match = re.search(r'ozon\.ru/product/(?:.*?)-?(\d+)', str(url))
            if match:
                product_id = match.group(1)
            else:
                match2 = re.search(r'ozon\.ru/product/(\d+)', str(url))
                if match2:
                    product_id = match2.group(1)
        
        if not product_id:
            # fallback to hash of url or something
            product_id = str(hash(name))
            
        seller = str(sheet.cell(row=row, column=4).value or "")
        brand = str(sheet.cell(row=row, column=5).value or "")
        
        size_text, length = parse_size(name)
        feet_cat = get_feet_category(length)
        
        revenue = safe_float(sheet.cell(row=row, column=11).value)
        sales = safe_int(sheet.cell(row=row, column=13).value)
        price = safe_float(sheet.cell(row=row, column=14).value)
        buyout_percent = safe_float(sheet.cell(row=row, column=16).value)
        
        impressions = safe_int(sheet.cell(row=row, column=24).value)
        card_views = safe_int(sheet.cell(row=row, column=26).value)
        
        ctr = 0.0
        if impressions > 0:
            ctr = round((card_views / impressions) * 100, 2)
            
        add_to_cart_percent = safe_float(sheet.cell(row=row, column=30).value)
        drr = safe_float(sheet.cell(row=row, column=35).value)
        
        product = {
            "id": product_id,
            "name": name,
            "url": str(url),
            "seller": seller,
            "brand": brand,
            "sizeText": size_text,
            "feetCategory": feet_cat,
            "revenue": revenue,
            "sales": sales,
            "price": price,
            "buyoutPercent": buyout_percent,
            "funnel": {
                "impressions": impressions,
                "cardViews": card_views,
                "ctr": ctr,
                "addToCartPercent": add_to_cart_percent,
                "drr": drr
            }
        }
        products.append(product)
        
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
        
    print(f"Processed {len(products)} products. Output saved to {output_file}")

if __name__ == "__main__":
    process_file()
