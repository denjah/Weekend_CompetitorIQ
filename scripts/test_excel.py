import openpyxl
import sys
import json
import codecs

# Use utf-8 for print
sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

file_path = r'd:\!PROJECTS\CONCURENTS\MANAGER\Аэрохоккей за 28 дней от 26.05 (2).xlsx'
print("Checking file:", file_path)
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
print("Headers:")
print(headers)

file_path2 = r'd:\!PROJECTS\CONCURENTS\MANAGER\ТЗ на обработку выгрузки.xlsx'
print("\nChecking file:", file_path2)
wb2 = openpyxl.load_workbook(file_path2, data_only=True)
print("Sheets:", wb2.sheetnames)
for sheet_name in wb2.sheetnames:
    sheet2 = wb2[sheet_name]
    headers2 = [cell.value for cell in sheet2[1]]
    row2 = [cell.value for cell in sheet2[2]]
    print(f"Sheet '{sheet_name}' Headers:", headers2)
    row3 = [cell.value for cell in sheet2[3]]
    row4 = [cell.value for cell in sheet2[4]]
    print(f"Sheet '{sheet_name}' Row3:", row3)
    print(f"Sheet '{sheet_name}' Row4:", row4)
    print(f"Sheet '{sheet_name}' Max Row:", sheet2.max_row)

print("\nChecking file 1 again:")
sheet1 = wb.active
print(f"Sheet 1 Max Row:", sheet1.max_row)
row3_1 = [cell.value for cell in sheet1[3]]
row4_1 = [cell.value for cell in sheet1[4]]
print("Sheet 1 Row3:", row3_1)
print("Sheet 1 Row4:", row4_1)
